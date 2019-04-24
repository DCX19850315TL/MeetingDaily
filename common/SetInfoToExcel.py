#!/usr/bin/env python
#_*_ coding:utf-8 _*_
'''
@author: tanglei
@contact: tanglei_0315@163.com
@file: SetInfoToExcel.py
@time: 2019/1/14 11:26
'''
#导入第三方模块
from openpyxl import Workbook
import os
import time
import shutil
#导入自定义的模块
from common.GetMeetingInfo import GetMeetingInfo,GetDetailedInformation,GetNubeInfoApiParams
from common.GetMeetingInfo import GetMeetingInfoApi,GetInfoApiParams,Headers,LoginApiParams,MeetingLoginApi
from common.GetUserInfo import GetUserAll
from common.logger import logger
#设置当前时间的变量
time = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime())
#设定上一层目录
'''
path = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
excel_file = os.path.join(path,"MeetingDaily.xlsx")
excel_backup_file = os.path.join(path,"excel_backup\MeetingDaily_%s.xlsx" %(time))
nube_excel_file = os.path.join(path,'NubeMeetingDaily.xlsx')
nube_backup_excel_file = os.path.join(path,'excel_backup\NubeMeetingDaily_%s.xlsx' % (time))
'''
excel_file = os.path.join(os.path.abspath(''),'MeetingDaily.xlsx')
excel_backup_file = os.path.join(os.path.abspath('excel_backup'),'MeetingDaily_%s.xlsx' % (time))
nube_excel_file = os.path.join(os.path.abspath(''),'NubeMeetingDaily.xlsx')
nube_backup_excel_file = os.path.join(os.path.abspath('excel_backup'),'NubeMeetingDaily_%s.xlsx' % (time))
MeetingApi = GetMeetingInfo(MeetingLoginApi, LoginApiParams, Headers)
MeetingApi.Login()
MeetingApi = GetMeetingInfo(GetMeetingInfoApi, GetInfoApiParams, Headers)
MeetingApi.GetInfo()
MeetingInfo = GetDetailedInformation(MeetingApi.GetInfo())
MeetingApi = GetMeetingInfo(GetMeetingInfoApi, GetNubeInfoApiParams(), Headers)
NubeMeetingInfo = MeetingApi.GetNubeInfo()

def SetInfoToExcel():
    # 会议个数
    MeetingLen = MeetingInfo.List_len()
    if os.path.isfile(excel_file):
        shutil.move(excel_file,excel_backup_file)
    Header = {"A1":"序号","B1":"会议号","C1":"会议时间","D1":"持续时间","E1":"端到端总体合格率","F1":"平均合格率","G1":"端到端最高合格率","H1":"端到端最低合格率","I1":"端到端|合格状态|丢包率|空音包","J1":"开会方数","K1":"用户列表","L1":"设备类型"}
    wb = Workbook()
    ws = wb.active
    for k,v in Header.items():
        ws[k] = v
        wb.save(excel_file)
    for item in range(MeetingLen):
        Body = {"A%d" % (item + 2): item + 1,
                "B%d" % (item + 2): MeetingInfo.GetMeetingNumber()[item],
                "C%d" % (item + 2): MeetingInfo.GetMeetingTime()[item],
                "D%d" % (item + 2): MeetingInfo.GetMeetingDuration()[item],
                "E%d" % (item + 2): MeetingInfo.GetPtoPTotal()[item],
                "F%d" % (item + 2): "丢包:%s\n空音包:%s" % (
            MeetingInfo.PacketLossAverage()[item], MeetingInfo.SoundPacketLossAverage()[item]),
                "G%d" % (item + 2): "丢包:%s\n空音包:%s" % (
                MeetingInfo.PacketLossMax()[item], MeetingInfo.SoundPacketLossMax()[item]),
                "H%d" % (item + 2): "丢包:%s\n空音包:%s" % (
                MeetingInfo.PacketLossMin()[item], MeetingInfo.SoundPacketLossMin()[item]),
                "I%d" % (item + 2): "\n".join(MeetingInfo.DetailData()[item]),
                "J%d" % (item + 2): MeetingInfo.MeetingPeopleNumber()[item],
                "K%d" % (item + 2): "\n".join(GetUserAll()[0][item]),
                "L%d" % (item + 2): "\n".join(GetUserAll()[1][item])}
        for k,v in Body.items():
            ws[k] = v
            wb.save(excel_file)

def SetNubeInfoToExcel():
    NubeLen = len(NubeMeetingInfo)
    if os.path.isfile(nube_excel_file):
        shutil.move(nube_excel_file,nube_backup_excel_file)
    Header = {"A1": "序号", "B1": "会议号", "C1": "会议开始时间", "D1": "持续时间", "E1": "人数", "F1": "用户列表"}
    wb = Workbook()
    ws = wb.active
    for k, v in Header.items():
        ws[k] = v
        wb.save(nube_excel_file)
    isNubeList = []
    for item in range(NubeLen):
        #判断数据中是否有items字段
        if "items" in NubeMeetingInfo[item]:
            isNubeList.append(NubeMeetingInfo[item])
    k = 0
    usercount_list = []
    for item in range(len(isNubeList)):
        isNubeListMoreInfo = isNubeList[item]["items"]
        for j in range(len(isNubeListMoreInfo)):
            ws["A%d" % (k+2)] = k+1
            ws["B%d" % (k+2)] = isNubeListMoreInfo[j]["meetingId"]
            ws["C%d" % (k + 2)] = isNubeListMoreInfo[j]["timeStamps"]
            time_seconds = isNubeListMoreInfo[j]["duration"]
            m,s = divmod(time_seconds,60)
            h,m = divmod(m,60)
            ws["D%d" % (k + 2)] = ("%02d:%02d:%02d" % (h, m, s))
            for kk in isNubeListMoreInfo[j]["userIdList"]:
                if len(kk) == 8:
                    usercount_list.append(kk)
            ws["E%d" % (k + 2)] = len(usercount_list)
            ws["F%d" % (k + 2)] = "\n".join(usercount_list)
            wb.save(nube_excel_file)
            k = k+1
            usercount_list = []